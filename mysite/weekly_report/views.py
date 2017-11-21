# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render
from django.http import HttpResponse
from django.http import StreamingHttpResponse
from django.template import Context,Template

import requests
from bs4 import BeautifulSoup

from itertools import product
import types

import openpyxl
from openpyxl import worksheet
from openpyxl import load_workbook
from openpyxl import Workbook

import time
import datetime
from datetime import timedelta

import re
# Create your views here.

def date_time():
    time_today = datetime.date.today()
    time1 = timedelta(days=7)
    time2 = timedelta(days=2)
    time_ago = time_today - time1
    time_now = time_today - time2
    _times = '%s' ' ---- ' '%s' % (time_ago, time_now)
    dic = {"_times":_times,"time_ago":time_ago,"time_now":time_now}
    return dic




#登录部分
root_url = 'http://172.16.203.12/zentao/user-login.html'

my_url = 'http://172.16.203.12/zentao/my-project.html'
r = requests.Session()
UA = "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36"

header = {"User-Agent": UA,
           "referer":"http://172.16.203.12/zentao/my/"
           }
f = r.get(root_url,headers = header)

#cookie登录
r.cookies = requests.utils.cookiejar_from_dict({
    
    'zentaosid':'u7p7ag69deghi44hh9r6au97t0'})
r.post(root_url,
    cookies = r.cookies,
    headers = header
    )
#账号密码登录
'''
postdata = {
    'account':'baozhong',
    'password':'111111'
}
r.post(
    root_url,
    data = postdata,
    headers = header)
'''

#抓取迭代名称
h = r.get(my_url,headers = header)
_soup = BeautifulSoup(h.content,'lxml')
diedais = _soup.select("table tbody tr")
task_names = []             #task_names是迭代名称
ID_names = []               #ID_names是迭代ID
for diedai in diedais:
    lt = list(diedai.stripped_strings)
    task_names.append(lt[2])
    ID_names.append(lt[0])

planA1_names = []               #迭代名称
planA2_names = []
out_plan_names = []

A1_IDs = []                    #IDs是迭代的编号
A2_IDs = []
out_IDs = []
IDs = []

A1_ID_names = []                #A1的ID
A2_ID_names = []                #A2的ID
out_plan_ID_names = []          #计划外任务的ID
for n,k in zip(task_names,ID_names):
    m = n[:5]                                           #取迭代名称的前5个字符串   
    x = re.match(r'^(\d{1}).(\d{1}).(\d{1})$', m)       #判断前5个字符串是不是数字（1.2.3类型的）
    _m = n[:6]
    y = re.match(r'^(\d{1}).(\d{1}).(\d{2})$', _m)
    if y:
        IDs.append(_m)
        if x.group(1) == '1':
            planA1_names.append(n[7:])
            A1_IDs.append(_m)
            A1_ID_names.append(k)
        elif x.group(1) == '2':
            planA2_names.append(n[7:])
            A2_IDs.append(_m)
            A1_ID_names.append(k)
    elif x:
        IDs.append(m)
        if x.group(1) == '1':       #如果第一个.之后是1，就是planA1计划的任务
            planA1_names.append(n[6:])
            A1_IDs.append(m)
            A1_ID_names.append(k)
        elif x.group(1) == '2':       #如果第一个.之后是2，就是planA2计划的任务
            planA2_names.append(n[6:])
            A2_IDs.append(m)
            A2_ID_names.append(k)
    
    else:
        out_IDs.append('')
        out_plan_names.append(n)
        out_plan_ID_names.append(k)


def get_data(ID_names,task_names,IDs):                #IDnames是迭代ID，IDs是迭代的编号，task_names是迭代名称
#抓取任务名称
                
    d = {}                                            #d是含迭代名称和对应的任务名称的字典
    dic = {}                                          #dic是迭代名称和迭代编号对应的字典
    next_contents = []
    for ID_name,task_name,ID in zip(ID_names,task_names,IDs):
        index_url = 'http://172.16.203.12/zentao/project-task-' + ID_name + '.html'
        f = r.get(index_url,headers = header)
        soup = BeautifulSoup(f.content,'lxml')
        plans = soup.select("table tbody tr")  
        contents = []
        for plan in plans:
            l = list(plan.stripped_strings)
            task_ID_name = l[0]                     #task_ID_name 是任务名称ID
            task_url = 'http://172.16.203.12/zentao/task-view-' + task_ID_name + '.html'     #这是某具体任务名称的网页地址
            #进入某具体任务名称网页抓取该任务名称的历史记录时间
            t = r.get(task_url,headers = header)
            t_soup = BeautifulSoup(t.content,'lxml')
            time_logs = t_soup.find_all('span', class_ = "item")
            time_logs_end = t_soup.find_all('tr')

            b_times = date_time()
            time_ago = b_times['time_ago']
            time_now = b_times['time_now']
#抓取任务历史记录时间
            for time_log in time_logs:
                lis = list(time_log.stripped_strings)
                c = lis[0]
                _lis = c[:10]                               #_lis是一个字符串历史记录时间               
                a = time.strptime(_lis,"%Y-%m-%d")
                b = datetime.date(*a[:3])                  #转换成时间格式的历史记录时间
           
                if b >= time_ago and b <= time_now:
                    if l[-1][-1:] == '%' and l[1].isdigit() == True:
                        contents.append(dict(name = l[2],jindu = l[-1]))
                        d[task_name] = contents
                        dic[task_name] = ID
                    elif l[-1][-1:] == '%':
                        contents.append(dict(name = l[1],jindu = l[-1]))
                        d[task_name] = contents
                        dic[task_name] = ID
                    else:
                        contents.append(dict(name = l[1],jindu = l[-2]))
                        d[task_name] = contents
                        dic[task_name] = ID
                    break
                else:
                    pass
#抓取任务截止日期和开始日期
            for time_ends in time_logs_end:                 
                time_a1 = '2000-01-01'
                time_b1 = '2000-01-01'
                if time_ends.th.string == u'截止日期':
                    time_a = list(time_ends.td.stripped_strings)
                    if time_a[0] == u'0000-00-00':
                        pass
                    else:
                        time_a1 = str(time_a[0])
                elif time_ends.th.string == u'实际开始':
                    time_c = list(time_ends.td.stripped_strings)
                    if time_c[0] == u'0000-00-00':
                        pass
                    else: 
                        time_b1 = str(time_c[0])
                else:
                    pass
                time_b = time.strptime(time_a1,"%Y-%m-%d")
                time_end = datetime.date(*time_b[:3])

                time_d = time.strptime(time_b1,"%Y-%m-%d")
                time_start = datetime.date(*time_d[:3])
                if ((time_start >= time_ago and time_start <= time_now) or time_start <= time_ago) and ((time_end >= time_ago and time_end <= time_now) or time_end >= time_now):
                    if l[1].isdigit() == True and l[-1][-1:] == '%':
                        if l[3] == '进行中':
                            next_contents.append(dict(task = l[2],people = l[-5]))
                        else:
                            pass

                    elif l[1].isdigit() == True and l[-1][-1:] != '%':
                        if l[3] == '进行中':
                            next_contents.append(dict(task = l[2],people = l[-6]))
                        else:
                            pass
                    else:
                        if l[2] == '进行中':
                            next_contents.append(dict(task = l[1],people = l[-5]))
                        else:
                            pass
                else:
                    pass

    for q in next_contents:
        print q['people'],q['task']


    ret = {'d':d,'dic':dic,'next_contents':next_contents}
    return ret 

    
        



data_A1 = get_data(A1_ID_names,planA1_names,A1_IDs)
print '-------------------'
data_A2 = get_data(A2_ID_names,planA2_names,A2_IDs)
print '-------------------'
data_out_plan = get_data(out_plan_ID_names,out_plan_names,out_IDs)

data_all = get_data(ID_names,task_names,IDs)



def index(request):
    times = date_time()
    time = times['_times']
    data_A1_d = data_A1['d']
    data_A2_d = data_A2['d']
    data_out_plan_d = data_out_plan['d']
    data_next = data_all['next_contents']
    context = {
    'time':time,
    'data_A1_d':data_A1_d,
    'data_A2_d':data_A2_d,
    'data_out_plan_d':data_out_plan_d,
    'data_next':data_next
        }

    return render(request,'weekly_report/index.html',context)



def downloadFile(request):
    times = date_time()
    time = times['_times']

    
    wb = load_workbook("zhoubao.xlsx")
    ws = wb.active
    num1 = 9
    for planA1_name,A1_ID in zip(data_A1['d'],data_A1['dic']):
        num1 = num1 + 1
        _cell1 = 'B' + str(num1)
        ws[_cell1] = planA1_name
        cell1_ = 'D' + str(num1)
        ws[cell1_] = data_A1['dic'][A1_ID]
        cell1 = 'E' + str(num1)
        x1 = ''
        for i1 in data_A1['d'][planA1_name]:
            x1 += i1['name'] + '' + i1['jindu'] + '\n'
        ws[cell1] = x1

    num2 = 12
    for planA2_name,A2_ID in zip(data_A2['d'],data_A2['dic']):
        num2 = num2 + 1
        _cell2 = 'B' + str(num2)
        ws[_cell2] = planA2_name
        cell2_ = 'D' + str(num2)
        ws[cell2_] = data_A2['dic'][A2_ID]
        cell2 = 'E' + str(num2)
        x2 = ''
        for i2 in data_A2['d'][planA2_name]:
            x2 += i2['name'] + '' + i2['jindu'] + '\n'
        ws[cell2] = x2

    num3 = 21
    for out_plan_name in data_out_plan['d']:
        num3 = num3 + 1
        _cell3 = 'B' + str(num3)
        ws[_cell3] = out_plan_name
        cell3 = 'E' + str(num3)
        x3 = ''
        for i3 in data_out_plan['d'][out_plan_name]:
            x3 += i3['name'] + '' + i3['jindu'] + '\n'
        ws[cell3] = x3

    num4 = 33
    for i in data_all['next_contents']:
        num4 += 1
        _cell4 = 'A' + str(num4)
        ws[_cell4] = i['task']
        cell4 = 'F' + str(num4)
        ws[cell4] = i['people']

    ws['G7'] = time
    wb.save('new_zhoubao.xlsx')
    file_name = 'new_zhoubao.xlsx'
    def file_iterator(file_name, chunk_size=512):#用于形成二进制数据  
        with open(file_name,'rb') as f:  
            while True:  
                c = f.read(chunk_size)
                if c:  
                    yield c
                else:
                    break  
    the_file_name ="new_zhoubao.xlsx"#要下载的文件路径  
    response =StreamingHttpResponse(file_iterator(the_file_name))#这里创建返回  
    response['Content-Type'] = 'application/vnd.ms-excel'#注意格式   
    response['Content-Disposition'] = 'attachment;filename="zhoubao.xlsx"'#注意filename 这个是下载后的名字  
    return response


#这是一个解决openpyxl打开表格部分边框缺失的补丁
def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self._merged_cells:
            self._merged_cells.append(range_string)


        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 1, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    m = types.MethodType(merge_cells, None, worksheet.Worksheet)
    worksheet.Worksheet.merge_cells = m


patch_worksheet()

